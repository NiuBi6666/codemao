import os

from openpyxl import load_workbook

from .db import get_db, utcnow


EXPECTED_HEADERS = ("用户ID", "姓名", "性别", "年龄", "年级", "班级名称")


class RosterValidationError(ValueError):
    pass


def _text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_roster(file_or_stream):
    try:
        workbook = load_workbook(file_or_stream, read_only=True, data_only=True)
    except Exception as exc:
        raise RosterValidationError("无法读取 Excel，请确认文件为有效的 .xlsx 格式") from exc

    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = tuple(_text(value) for value in next(iterator))
        except StopIteration as exc:
            raise RosterValidationError("Excel 内容为空") from exc

        if headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
            raise RosterValidationError("表头必须依次为：" + "、".join(EXPECTED_HEADERS))

        rows = []
        user_ids = set()
        for excel_row, values in enumerate(iterator, start=2):
            cells = tuple(_text(value) for value in values[: len(EXPECTED_HEADERS)])
            if not any(cells):
                continue
            user_id, name, gender, age, grade, class_name = cells
            if not user_id or not name:
                raise RosterValidationError(f"第 {excel_row} 行缺少用户ID或姓名")
            if user_id in user_ids:
                raise RosterValidationError(f"第 {excel_row} 行用户ID重复：{user_id}")
            user_ids.add(user_id)
            rows.append(
                {
                    "user_id": user_id,
                    "name": name,
                    "gender": gender,
                    "age": age,
                    "grade": grade,
                    "class_name": class_name,
                }
            )
        if not rows:
            raise RosterValidationError("Excel 中没有可导入的学生数据")
        if len(rows) > 100000:
            raise RosterValidationError("单次导入不能超过 100000 行")
        return rows
    finally:
        workbook.close()


def replace_roster(rows, filename, imported_by):
    connection = get_db()
    imported_at = utcnow()
    try:
        connection.execute("BEGIN IMMEDIATE")
        connection.execute("DELETE FROM students")
        connection.executemany(
            """
            INSERT INTO students (user_id, name, gender, age, grade, class_name, imported_at)
            VALUES (:user_id, :name, :gender, :age, :grade, :class_name, :imported_at)
            """,
            ({**row, "imported_at": imported_at} for row in rows),
        )
        connection.execute(
            "INSERT INTO import_batches (filename, row_count, imported_by, created_at) VALUES (?, ?, ?, ?)",
            (filename, len(rows), imported_by, imported_at),
        )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    return len(rows)


def import_roster_path(path, imported_by):
    rows = parse_roster(path)
    return replace_roster(rows, os.path.basename(path), imported_by)
